from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import io
import pandas as pd
from fastapi import Form
from collections import defaultdict

from database.db import get_db
from database.tables.accounts import Account
from database.tables.users import User
from database.tables.targets import Target, TargetCreate, AccountTypeEnum
from database.tables.actuals import Actual
from api.v1.endpoints.auth import get_current_user
from services.access_scope_service import is_admin_user, is_sales_user

router = APIRouter()

class DownloadTemplateRequest(BaseModel):
    account_ids: List[int]
    ecode: str

class UploadResponse(BaseModel):
    message: str
    total_records: int
    success_count: int
    errors: List[str] = []


def normalize_fy_value(raw_fy: str) -> Optional[str]:
    value = str(raw_fy or "").strip()
    if not value:
        return None

    parts = value.replace("/", "-").split("-")
    if len(parts) != 2:
        return None

    left = parts[0].strip()
    right = parts[1].strip()
    if not left.isdigit() or not right.isdigit():
        return None

    if len(left) == 2:
        start_year = 2000 + int(left)
    else:
        start_year = int(left)

    if len(right) == 2:
        end_year = 2000 + int(right)
    else:
        end_year = int(right)

    if end_year < start_year:
        return None

    return f"{start_year}-{end_year}"


def current_financial_year() -> str:
    now = datetime.now()
    start_year = now.year if now.month >= 4 else now.year - 1
    end_year = start_year + 1
    return f"{start_year}-{end_year}"

@router.get("/financial-years")
async def get_financial_years(
    ecode: str = Query(None, description="ECode to get financial years for. If not provided, returns all financial years"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get all distinct financial years for a specific ECode, or all financial years if no ECode provided
    """
    try:
        if is_sales_user(current_user) and not is_admin_user(current_user):
            if ecode and ecode != current_user.ECode:
                raise HTTPException(status_code=403, detail="Sales user can only access own financial years")
            ecode = current_user.ECode

        if ecode:
            # Validate ECode exists
            user = db.query(User).filter(User.ECode == ecode).first()
            if not user:
                raise HTTPException(status_code=404, detail=f"User with ECode '{ecode}' not found")
            
            # Get distinct FY values for this ECode
            financial_years = db.query(Target.FY).filter(Target.ECode == ecode).distinct().all()
        else:
            # Get all distinct FY values from both targets and actuals tables
            targets_fy = db.query(Target.FY).distinct()
            actuals_fy = db.query(Actual.FY).distinct()
            
            # Use UNION to combine both queries and get distinct FY values
            union_query = targets_fy.union(actuals_fy)
            financial_years = union_query.all()

        normalized = []
        for fy in financial_years:
            normalized_fy = normalize_fy_value(fy[0])
            if normalized_fy:
                normalized.append(normalized_fy)

        unique_sorted = sorted(set(normalized), key=lambda x: int(x.split("-")[0]), reverse=True)
        if not unique_sorted:
            unique_sorted = [current_financial_year()]

        return {"financial_years": unique_sorted}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching financial years: {str(e)}")

@router.get("/user-targets")
async def get_user_targets(
    ecode: str = Query(..., description="ECode to get targets for"),
    fy: str = Query(..., description="Financial year to filter by"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get targets for a specific user and financial year in the required format
    """
    try:
        if is_sales_user(current_user) and not is_admin_user(current_user) and ecode != current_user.ECode:
            raise HTTPException(status_code=403, detail="Sales user can only access own targets")

        # Validate ECode exists
        user = db.query(User).filter(User.ECode == ecode).first()
        if not user:
            raise HTTPException(status_code=404, detail=f"User with ECode '{ecode}' not found")
        
        # Get all targets for this ECode and FY with account information
        targets = db.query(Target).options(
            joinedload(Target.account)
        ).filter(
            Target.ECode == ecode,
            Target.FY == fy
        ).all()
        
        if not targets:
            return []
        
        # Group targets by account
        account_targets = defaultdict(lambda: {
            "Account": None,
            "AccountType": None,
            "Apr": None, "May": None, "Jun": None, "Jul": None,
            "Aug": None, "Sep": None, "Oct": None, "Nov": None,
            "Dec": None, "Jan": None, "Feb": None, "Mar": None
        })
        
        for target in targets:
            # Create account key - use AccountID for EXISTING, 'NEW' for NEW accounts
            if target.AccountType == AccountTypeEnum.NEW:
                account_key = "NEW"
                account_name = "NEW"
            else:
                account_key = target.AccountID
                if target.account:
                    # Format account name as {Name} - {Division} - {Location}
                    division = target.account.Division or ""
                    location = target.account.Location or ""
                    account_name = f"{target.account.Name}"
                    if division:
                        account_name += f" - {division}"
                    if location:
                        account_name += f" - {location}"
                else:
                    account_name = f"Unknown Account (ID: {target.AccountID})"
            
            # Set account name if not already set
            if account_targets[account_key]["Account"] is None:
                account_targets[account_key]["Account"] = account_name
                account_targets[account_key]["AccountType"] = target.AccountType.value
            
            # Set the target value for the specific month
            if target.Month in account_targets[account_key]:
                account_targets[account_key][target.Month] = target.Target
        
        # Convert to list format and remove AccountType from response
        result = []
        for account_data in account_targets.values():
            # Remove AccountType before returning (we only needed it for processing)
            account_result = {k: v for k, v in account_data.items() if k != "AccountType"}
            result.append(account_result)
        
        # Sort results: NEW accounts last, others by account name
        result.sort(key=lambda x: (x["Account"] == "NEW", x["Account"]))
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching user targets: {str(e)}")

@router.post("/download-template")
async def download_template(
    request: DownloadTemplateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Generate and download an Excel template for targets with specified accounts and salesperson
    """
    try:
        if is_sales_user(current_user) and not is_admin_user(current_user) and request.ecode != current_user.ECode:
            raise HTTPException(status_code=403, detail="Sales user can download template only for self")

        # Fetch salesperson information
        salesperson = db.query(User).filter(User.ECode == request.ecode).first()
        if not salesperson:
            raise HTTPException(status_code=404, detail=f"Salesperson with ECode '{request.ecode}' not found")
        
        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Targets Template"
        
        # Add salesperson name at the top (row 1)
        salesperson_cell = ws.cell(row=1, column=1, value=f"{salesperson.Name}")
        salesperson_cell.font = Font(bold=True, size=14)
        
        # Define the headers (now in row 3)
        headers = ["Account Name", "FY", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        
        # Write headers to the third row (leaving row 2 blank for spacing)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Fetch accounts based on provided IDs
        if request.account_ids:
            accounts = db.query(Account).filter(Account.id.in_(request.account_ids)).all()
        else:
            accounts = []
        
        # Populate account rows (starting from row 4)
        row_num = 4
        for account in accounts:
            # Format account name as {Name} - {Division} - {Location}
            division = account.Division or ""
            location = account.Location or ""
            account_name = f"{account.Name}"
            if division:
                account_name += f" - {division}"
            if location:
                account_name += f" - {location}"
            
            ws.cell(row=row_num, column=1, value=account_name)
            row_num += 1
        
        # Add "NEW" row at the end
        ws.cell(row=row_num, column=1, value="NEW")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create streaming response
        response = StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=targets_template_{salesperson.Name.replace(' ', '_')}.xlsx"}
        )
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating template: {str(e)}")

@router.post("/upload-template", response_model=UploadResponse)
async def upload_template(
    file: UploadFile = File(...),
    ecode: str = Form(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Upload and process an Excel template file to populate targets data
    """
    try:
        if is_sales_user(current_user) and not is_admin_user(current_user) and ecode != current_user.ECode:
            raise HTTPException(status_code=403, detail="Sales user can upload targets only for self")

        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
        
        # Read the Excel file
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), header=2)  # Header starts at row 3 (0-indexed as 2)
        
        # Validate required columns
        expected_columns = ["Account Name", "FY", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        missing_columns = [col for col in expected_columns if col not in df.columns]
        
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        if not ecode:
            raise HTTPException(status_code=400, detail="ECode must be provided")
        
        # Validate ECode exists
        user = db.query(User).filter(User.ECode == ecode).first()
        if not user:
            raise HTTPException(status_code=404, detail=f"User with ECode '{ecode}' not found")
        
        # Process data
        success_count = 0
        errors = []
        month_columns = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        
        for index, row in df.iterrows():
            account_name = str(row["Account Name"]).strip()
            fy = str(row["FY"]).strip() if pd.notna(row["FY"]) else ""
            
            # Skip empty rows
            if pd.isna(row["Account Name"]) or account_name == "":
                continue
            
            # Determine account type and ID
            account_type = AccountTypeEnum.NEW if account_name.upper() == "NEW" else AccountTypeEnum.EXISTING
            account_id = None
            
            if account_type == AccountTypeEnum.EXISTING:
                # Find account by name (handle the formatted name)
                account = None
                accounts = db.query(Account).all()
                
                for acc in accounts:
                    # Create formatted name to match
                    division = acc.Division or ""
                    location = acc.Location or ""
                    formatted_name = f"{acc.Name}"
                    if division:
                        formatted_name += f" - {division}"
                    if location:
                        formatted_name += f" - {location}"
                    
                    if formatted_name == account_name or acc.Name == account_name:
                        account = acc
                        account_id = acc.id
                        break
                
                if not account:
                    errors.append(f"Row {index + 4}: Account '{account_name}' not found")
                    continue
            
            if not fy:
                errors.append(f"Row {index + 4}: FY is required")
                continue
            
            # Process each month
            for month in month_columns:
                target_value = row[month]
                
                # Skip if target value is empty or NaN
                if pd.isna(target_value) or str(target_value).strip() == "":
                    continue
                
                try:
                    target_value = float(target_value)
                except (ValueError, TypeError):
                    errors.append(f"Row {index + 4}, {month}: Invalid target value '{target_value}'")
                    continue
                
                # Check if record already exists
                existing_target = db.query(Target).filter(
                    Target.ECode == ecode,
                    Target.AccountID == account_id,
                    Target.AccountType == account_type,
                    Target.FY == fy,
                    Target.Month == month
                ).first()
                
                if existing_target:
                    # Update existing record
                    existing_target.Target = target_value
                else:
                    # Create new record
                    new_target = Target(
                        ECode=ecode,
                        AccountID=account_id,
                        AccountType=account_type,
                        FY=fy,
                        Month=month,
                        Target=target_value
                    )
                    db.add(new_target)
                
                success_count += 1
        
        # Commit all changes
        db.commit()
        
        return UploadResponse(
            message="File processed successfully",
            total_records=len(df),
            success_count=success_count,
            errors=errors
        )
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@router.get("/", response_model=List[dict])
async def get_targets(
    ecode: str = None,
    account_id: int = None,
    account_type: str = None,
    fy: str = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get targets with optional filters
    """
    query = db.query(Target)
    
    if is_sales_user(current_user) and not is_admin_user(current_user):
        query = query.filter(Target.ECode == current_user.ECode)
    elif ecode:
        query = query.filter(Target.ECode == ecode)
    if account_id:
        query = query.filter(Target.AccountID == account_id)
    if account_type:
        query = query.filter(Target.AccountType == account_type)
    if fy:
        query = query.filter(Target.FY == fy)
    
    targets = query.all()
    
    # Format response with related data
    result = []
    for target in targets:
        result.append({
            "ID": target.ID,
            "ECode": target.ECode,
            "UserName": target.user.Name if target.user else None,
            "AccountID": target.AccountID,
            "AccountName": target.account.Name if target.account else None,
            "AccountType": target.AccountType.value,
            "FY": target.FY,
            "Month": target.Month,
            "Target": target.Target
        })
    
    return result
