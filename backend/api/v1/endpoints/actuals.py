from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import text
from typing import List
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import io
import pandas as pd

from database.db import get_db
from database.tables.users import User
from database.tables.accounts import Account
from database.tables.deals import Deal
from database.tables.targets import Target
from database.tables.actuals import Actual, ActualCreate, AccountTypeEnum
from api.v1.endpoints.auth import get_current_user
from services.access_scope_service import is_admin_user, is_sales_user

# Import configuration for KAM ratios
import sys
import os

from config import KAM_RATIO, SALESPERSON_RATIO

router = APIRouter()

class UploadResponse(BaseModel):
    message: str
    total_records: int
    success_count: int
    errors: List[str] = []

@router.post("/download-template")
async def download_template(
    fy: str = Query(..., description="Financial year (e.g., '25-26')"),
    month: str = Query(..., description="Month (Apr-Mar)"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Generate and download an Excel template for actuals with two sheets:
    1. Key Account Business
    2. New Business
    """
    try:
        if is_sales_user(current_user) and not is_admin_user(current_user):
            raise HTTPException(status_code=403, detail="Sales user cannot download full actuals template")

        # Validate month
        valid_months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        if month not in valid_months:
            raise HTTPException(status_code=400, detail=f"Invalid month. Must be one of: {', '.join(valid_months)}")
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet and create our custom sheets
        wb.remove(wb.active)
        
        # Create Sheet 1: Key Account Business
        ws1 = wb.create_sheet("Key Account Business")
        await create_key_account_business_sheet(ws1, fy, month, db)
        
        # Create Sheet 2: New Business
        ws2 = wb.create_sheet("New Business")
        await create_new_business_sheet(ws2, fy, month, db)
        
        # Save to BytesIO buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create streaming response
        response = StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=actuals_template_{fy}_{month}.xlsx"}
        )
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating template: {str(e)}")

async def create_key_account_business_sheet(ws, fy: str, month: str, db: Session):
    """Create the Key Account Business sheet"""
    
    # Set up styling
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title and FY/Month info
    ws.cell(row=1, column=1, value="Key Account Business").font = title_font
    ws.cell(row=2, column=1, value=f"FY: {fy}").font = header_font
    ws.cell(row=2, column=2, value=f"Month: {month}").font = header_font
    
    # Headers starting from row 4
    headers = ["ECode", "Salesperson Name", "Account Name", "Target", "Actuals (Excl. New Business)", "Qualified"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Execute query to get data
    query = text("""
        SELECT t.ECode, u.Name, a.Name, a.Division, a.Location, t.Target 
        FROM Targets t, Accounts a, Users u 
        WHERE a.ID = t.AccountID 
        AND u.ECode = t.ECode 
        AND t.AccountType = 'EXISTING' 
        AND t.FY = :fy 
        AND t.Month = :month
        ORDER BY u.Name, a.Name
    """)
    
    result = db.execute(query, {"fy": fy, "month": month})
    data = result.fetchall()
    
    # Populate data starting from row 5
    row_num = 5
    for row_data in data:
        ws.cell(row=row_num, column=1, value=row_data[0]).border = border  # ECode
        ws.cell(row=row_num, column=2, value=row_data[1]).border = border  # Salesperson Name
        # Format account name as {Name} - {Division} - {Location}
        division = row_data[3] or ""
        location = row_data[4] or ""
        account_name = f"{row_data[2]}"
        if division:
            account_name += f" - {division}"
        if location:
            account_name += f" - {location}"
        ws.cell(row=row_num, column=3, value=account_name).border = border  # Account Name
        ws.cell(row=row_num, column=4, value=row_data[5]).border = border  # Target
        ws.cell(row=row_num, column=5, value="").border = border  # Actuals (blank)
        ws.cell(row=row_num, column=6, value="").border = border  # Qualified (blank)
        row_num += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

@router.post("/upload-template", response_model=UploadResponse)
async def upload_template(
    file: UploadFile = File(...),
    fy: str = Form(...),
    month: str = Form(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Upload and process an Excel template file to populate actuals data
    """
    try:
        if is_sales_user(current_user) and not is_admin_user(current_user):
            raise HTTPException(status_code=403, detail="Sales user cannot upload full actuals template")

        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
        
        # Validate month
        valid_months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        if month not in valid_months:
            raise HTTPException(status_code=400, detail=f"Invalid month. Must be one of: {', '.join(valid_months)}")
        
        # Read the Excel file
        content = await file.read()
        
        # Process both sheets
        success_count = 0
        errors = []
        total_records = 0
        
        # Process Sheet 1: Key Account Business
        try:
            df_key_accounts = pd.read_excel(io.BytesIO(content), sheet_name="Key Account Business", header=3)  # Header starts at row 4 (0-indexed as 3)
            key_success, key_errors, key_total = await process_key_account_business_sheet(df_key_accounts, fy, month, db)
            success_count += key_success
            errors.extend(key_errors)
            total_records += key_total
        except Exception as e:
            errors.append(f"Error processing Key Account Business sheet: {str(e)}")
        
        # Process Sheet 2: New Business
        try:
            new_success, new_errors, new_total = await process_new_business_sheet(content, fy, month, db)
            success_count += new_success
            errors.extend(new_errors)
            total_records += new_total
        except Exception as e:
            errors.append(f"Error processing New Business sheet: {str(e)}")
        
        # Note: Each sheet processor handles its own atomic transactions
        
        return UploadResponse(
            message="File processed successfully",
            total_records=total_records,
            success_count=success_count,
            errors=errors
        )
        
    except HTTPException:
        # No need to rollback since sheet processors handle their own transactions
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

async def process_key_account_business_sheet(df, fy: str, month: str, db: Session):
    """Process the Key Account Business sheet"""
    success_count = 0
    errors = []
    total_records = len(df)
    
    # Expected columns for Key Account Business
    expected_columns = ["ECode", "Salesperson Name", "Account Name", "Target", "Actuals (Excl. New Business)", "Qualified"]
    missing_columns = [col for col in expected_columns if col not in df.columns]
    
    if missing_columns:
        errors.append(f"Key Account Business sheet missing columns: {', '.join(missing_columns)}")
        return 0, errors, 0
    
    # Create a savepoint for atomic operations
    savepoint = db.begin()
    
    try:
        for index, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row["ECode"]) or str(row["ECode"]).strip() == "":
                continue
            
            ecode = str(row["ECode"]).strip()
            account_name = str(row["Account Name"]).strip()
            actual_value = row["Actuals (Excl. New Business)"]
            qualified_value = row["Qualified"]
            
            # Validate ECode exists
            user = db.query(User).filter(User.ECode == ecode).first()
            if not user:
                errors.append(f"Key Account Business row {index + 5}: User with ECode '{ecode}' not found")
                continue
            
            # Find account by name
            account = None
            accounts = db.query(Account).all()
            
            for acc in accounts:
                # Create formatted name to match (same logic as in targets)
                division = acc.Division or ""
                location = acc.Location or ""
                formatted_name = f"{acc.Name}"
                if division:
                    formatted_name += f" - {division}"
                if location:
                    formatted_name += f" - {location}"
                
                if formatted_name == account_name or acc.Name == account_name:
                    account = acc
                    break
            
            if not account:
                errors.append(f"Key Account Business row {index + 5}: Account '{account_name}' not found")
                continue
            
            # Parse actual and qualified values
            actual_float = None
            qualified_float = None
            
            if not pd.isna(actual_value) and str(actual_value).strip() != "":
                try:
                    actual_float = float(actual_value)
                except (ValueError, TypeError):
                    errors.append(f"Key Account Business row {index + 5}: Invalid Actuals value '{actual_value}'")
                    continue
            
            if not pd.isna(qualified_value) and str(qualified_value).strip() != "":
                try:
                    qualified_float = float(qualified_value)
                except (ValueError, TypeError):
                    errors.append(f"Key Account Business row {index + 5}: Invalid Qualified value '{qualified_value}'")
                    continue
            
            # Check if record already exists
            existing_actual = db.query(Actual).filter(
                Actual.ECode == ecode,
                Actual.AccountID == account.id,
                Actual.AccountType == AccountTypeEnum.KAB,
                Actual.FY == fy,
                Actual.Month == month
            ).first()
            
            if existing_actual:
                # Update existing record
                existing_actual.Actual = actual_float
                existing_actual.Qualified = qualified_float
            else:
                # Create new record
                new_actual = Actual(
                    ECode=ecode,
                    AccountID=account.id,
                    AccountType=AccountTypeEnum.KAB,
                    FY=fy,
                    Month=month,
                    Actual=actual_float,
                    Qualified=qualified_float
                )
                db.add(new_actual)
            
            success_count += 1
    
    except Exception as e:
        # Rollback the transaction on any error
        savepoint.rollback()
        errors.append(f"Error processing Key Account Business sheet: {str(e)}")
        return 0, errors, total_records
    
    # If we have any errors, rollback the transaction
    if errors:
        savepoint.rollback()
        return 0, errors, total_records
    
    # If everything is successful, commit the transaction
    savepoint.commit()
    return success_count, errors, total_records

async def process_new_business_sheet(content, fy: str, month: str, db: Session):
    """Process the New Business sheet with its complex structure"""
    success_count = 0
    errors = []
    total_records = 0
    
    # Create a savepoint for atomic operations
    savepoint = db.begin()
    
    try:
        # Read the New Business sheet
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if "New Business" not in wb.sheetnames:
            errors.append("New Business sheet not found in workbook")
            return 0, errors, 0
        
        ws = wb["New Business"]
        
        current_row = 4  # Start from row 4 (after title and headers)
        current_entry = None  # Will store either single ECode or KAM pair info
        
        while current_row <= ws.max_row:
            cell_value = ws.cell(row=current_row, column=1).value
            
            if not cell_value:
                current_row += 1
                continue
            
            cell_str = str(cell_value)
            
            # Check if this is an employee header row (individual or KAM pair)
            if "ECode:" in cell_str:
                # Extract ECode(s) from the cell
                try:
                    if " & " in cell_str:
                        # This is a KAM pair: "SalespersonName & KAMName (ECode: SP001 & KAM001)"
                        names_part = cell_str.split("(ECode:")[0].strip()
                        ecode_part = cell_str.split("ECode:")[1].strip().rstrip(")")
                        
                        salesperson_name, kam_name = names_part.split(" & ")
                        salesperson_ecode, kam_ecode = ecode_part.split(" & ")
                        
                        # Validate both ECodes exist
                        salesperson = db.query(User).filter(User.ECode == salesperson_ecode.strip()).first()
                        kam = db.query(User).filter(User.ECode == kam_ecode.strip()).first()
                        
                        if not salesperson:
                            errors.append(f"New Business: Salesperson with ECode '{salesperson_ecode.strip()}' not found")
                            current_entry = None
                        elif not kam:
                            errors.append(f"New Business: KAM with ECode '{kam_ecode.strip()}' not found")
                            current_entry = None
                        else:
                            current_entry = {
                                "type": "kam_pair",
                                "salesperson_ecode": salesperson_ecode.strip(),
                                "kam_ecode": kam_ecode.strip()
                            }
                    else:
                        # This is a single salesperson: "Name (ECode: L4131)"
                        ecode_part = cell_str.split("ECode:")[1].strip().rstrip(")")
                        
                        # Validate ECode exists
                        user = db.query(User).filter(User.ECode == ecode_part).first()
                        if not user:
                            errors.append(f"New Business: User with ECode '{ecode_part}' not found")
                            current_entry = None
                        else:
                            current_entry = {
                                "type": "single",
                                "ecode": ecode_part
                            }
                    
                except Exception:
                    errors.append(f"New Business row {current_row}: Could not parse ECode from '{cell_str}'")
                    current_entry = None
                
                current_row += 1
                continue
            
            # Check if this is a target info row or KAM pair business label
            if "New Business Target:" in cell_str or cell_str == "KAM Pair Business":
                current_row += 1
                continue
            
            # Check if this is a table header row
            if cell_str == "Account Name":
                current_row += 1
                continue
            
            # This should be a data row if we have a current entry
            if current_entry and cell_str != "":
                total_records += 1
                
                account_name = cell_str
                type_value = ws.cell(row=current_row, column=2).value
                actual_value = ws.cell(row=current_row, column=3).value
                qualified_value = ws.cell(row=current_row, column=4).value
                
                # Determine AccountType
                account_type = AccountTypeEnum.NEW  # Default
                if type_value and str(type_value).strip().upper() == "EXISTING":
                    account_type = AccountTypeEnum.EXISTING
                elif type_value and str(type_value).strip().upper() == "NEW":
                    account_type = AccountTypeEnum.NEW
                
                # Find account by name (for both NEW and EXISTING types)
                account_id = None
                accounts = db.query(Account).all()
                for acc in accounts:
                    # Create formatted name to match
                    division = acc.Division or ""
                    location = acc.Location or ""
                    formatted_name = f"{acc.Name}"
                    if division:
                        formatted_name += f" - {division}"
                    if location:
                        formatted_name += f" - {location}"
                    
                    if formatted_name == account_name or acc.Name == account_name:
                        account_id = acc.id
                        break
                
                if not account_id:
                    errors.append(f"New Business row {current_row}: Account '{account_name}' not found in system")
                    current_row += 1
                    continue
                
                # Parse actual and qualified values
                actual_float = None
                qualified_float = None
                
                if actual_value is not None and str(actual_value).strip() != "":
                    try:
                        actual_float = float(actual_value)
                    except (ValueError, TypeError):
                        errors.append(f"New Business row {current_row}: Invalid Actuals value '{actual_value}'")
                        current_row += 1
                        continue
                
                if qualified_value is not None and str(qualified_value).strip() != "":
                    try:
                        qualified_float = float(qualified_value)
                    except (ValueError, TypeError):
                        errors.append(f"New Business row {current_row}: Invalid Qualified value '{qualified_value}'")
                        current_row += 1
                        continue
                
                # Skip if both actual and qualified are empty
                if actual_float is None and qualified_float is None:
                    current_row += 1
                    continue
                
                # Handle KAM pairs vs individual salesperson
                if current_entry["type"] == "kam_pair":
                    # Split values between KAM and Salesperson
                    kam_actual = actual_float * KAM_RATIO if actual_float is not None else None
                    kam_qualified = qualified_float * KAM_RATIO if qualified_float is not None else None
                    
                    salesperson_actual = actual_float * SALESPERSON_RATIO if actual_float is not None else None
                    salesperson_qualified = qualified_float * SALESPERSON_RATIO if qualified_float is not None else None
                    
                    # Create/update KAM record
                    existing_kam_actual = db.query(Actual).filter(
                        Actual.ECode == current_entry["kam_ecode"],
                        Actual.AccountID == account_id,
                        Actual.AccountType == account_type,
                        Actual.FY == fy,
                        Actual.Month == month
                    ).first()
                    
                    if existing_kam_actual:
                        existing_kam_actual.Actual = kam_actual
                        existing_kam_actual.Qualified = kam_qualified
                    else:
                        new_kam_actual = Actual(
                            ECode=current_entry["kam_ecode"],
                            AccountID=account_id,
                            AccountType=account_type,
                            FY=fy,
                            Month=month,
                            Actual=kam_actual,
                            Qualified=kam_qualified
                        )
                        db.add(new_kam_actual)
                    
                    # Create/update Salesperson record
                    existing_sp_actual = db.query(Actual).filter(
                        Actual.ECode == current_entry["salesperson_ecode"],
                        Actual.AccountID == account_id,
                        Actual.AccountType == account_type,
                        Actual.FY == fy,
                        Actual.Month == month
                    ).first()
                    
                    if existing_sp_actual:
                        existing_sp_actual.Actual = salesperson_actual
                        existing_sp_actual.Qualified = salesperson_qualified
                    else:
                        new_sp_actual = Actual(
                            ECode=current_entry["salesperson_ecode"],
                            AccountID=account_id,
                            AccountType=account_type,
                            FY=fy,
                            Month=month,
                            Actual=salesperson_actual,
                            Qualified=salesperson_qualified
                        )
                        db.add(new_sp_actual)
                    
                    success_count += 2  # Count both entries
                    
                else:
                    # Single salesperson - existing logic
                    existing_actual = db.query(Actual).filter(
                        Actual.ECode == current_entry["ecode"],
                        Actual.AccountID == account_id,
                        Actual.AccountType == account_type,
                        Actual.FY == fy,
                        Actual.Month == month
                    ).first()
                    
                    if existing_actual:
                        existing_actual.Actual = actual_float
                        existing_actual.Qualified = qualified_float
                    else:
                        new_actual = Actual(
                            ECode=current_entry["ecode"],
                            AccountID=account_id,
                            AccountType=account_type,
                            FY=fy,
                            Month=month,
                            Actual=actual_float,
                            Qualified=qualified_float
                        )
                        db.add(new_actual)
                    
                    success_count += 1
            
            current_row += 1
    
    except Exception as e:
        # Rollback the transaction on any error
        savepoint.rollback()
        errors.append(f"Error processing New Business sheet: {str(e)}")
        return 0, errors, total_records
    
    # If we have any errors, rollback the transaction
    if errors:
        savepoint.rollback()
        return 0, errors, total_records
    
    # If everything is successful, commit the transaction
    savepoint.commit()
    return success_count, errors, total_records

async def create_new_business_sheet(ws, fy: str, month: str, db: Session):
    """Create the New Business sheet with KAM pair handling"""
    
    # Set up styling
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    employee_font = Font(bold=True, size=13, color="4472C4")
    kam_pair_font = Font(bold=True, size=13, color="8B0000")  # Dark red for KAM pairs
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    employee_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    kam_pair_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title and FY/Month info
    ws.cell(row=1, column=1, value="New Business").font = title_font
    ws.cell(row=2, column=1, value=f"FY: {fy}").font = header_font
    ws.cell(row=2, column=2, value=f"Month: {month}").font = header_font
    
    # Get all users who have targets
    users_query = text("""
        SELECT DISTINCT u.ECode, u.Name 
        FROM Users u, Targets t 
        WHERE u.ECode = t.ECode 
        AND t.FY = :fy 
        AND t.Month = :month
        ORDER BY u.Name
    """)
    
    users_result = db.execute(users_query, {"fy": fy, "month": month})
    users = users_result.fetchall()
    
    current_row = 4
    
    # Track processed deals to avoid duplication
    processed_deals = set()
    
    for user_ecode, user_name in users:
        # Get new business target for this user
        target_query = text("""
            SELECT Target 
            FROM Targets 
            WHERE ECode = :ecode 
            AND AccountType = 'NEW' 
            AND FY = :fy 
            AND Month = :month
        """)
        
        target_result = db.execute(target_query, {"ecode": user_ecode, "fy": fy, "month": month})
        target_data = target_result.fetchone()
        target_value = target_data[0] if target_data else 0
        
        # Get accounts for this user - modified query to handle KAM deals
        # First, get individual salesperson deals (no KAM or same person as KAM)
        individual_accounts_query = text("""
            SELECT DISTINCT a.Name, a.Division, a.Location
            FROM Deals d
            JOIN Accounts a ON d.AccountID = a.ID
            WHERE d.SalespersonECode = :ecode
            AND d.Stage = 'DEAL_WON'
            AND (d.KAMECode IS NULL OR d.KAMECode = d.SalespersonECode)
            ORDER BY a.Name
        """)
        
        individual_accounts_result = db.execute(individual_accounts_query, {"ecode": user_ecode, "fy": fy, "month": month})
        individual_accounts = individual_accounts_result.fetchall()
        
        # Only show individual section if there are individual accounts
        if individual_accounts:
            # Employee header for individual
            employee_cell = ws.cell(row=current_row, column=1, value=f"{user_name} (ECode: {user_ecode})")
            employee_cell.font = employee_font
            employee_cell.fill = employee_fill
            ws.merge_cells(f"A{current_row}:D{current_row}")
            current_row += 1
            
            # Target info
            target_cell = ws.cell(row=current_row, column=1, value=f"New Business Target: {target_value}")
            target_cell.font = header_font
            ws.merge_cells(f"A{current_row}:D{current_row}")
            current_row += 1
            
            # Table headers
            table_headers = ["Account Name", "Customer Type", "Actuals", "Qualified"]
            for col_num, header in enumerate(table_headers, 1):
                cell = ws.cell(row=current_row, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            current_row += 1
            
            # Create dropdown validation for Customer Type column
            type_validation = DataValidation(
                type="list",
                formula1='"NEW,EXISTING"',
                allow_blank=True
            )
            
            # Add individual account rows
            start_data_row = current_row
            for account_name, division, location in individual_accounts:
                # Format account name
                formatted_name = account_name
                if division:
                    formatted_name += f" - {division}"
                if location:
                    formatted_name += f" - {location}"
                
                ws.cell(row=current_row, column=1, value=formatted_name).border = border
                type_cell = ws.cell(row=current_row, column=2, value="")
                type_cell.border = border
                ws.cell(row=current_row, column=3, value="").border = border
                ws.cell(row=current_row, column=4, value="").border = border
                current_row += 1
            
            # Apply dropdown validation to TYPE column for this user's section
            end_data_row = current_row - 1
            if end_data_row >= start_data_row:
                type_validation.add(f"B{start_data_row}:B{end_data_row}")
                ws.add_data_validation(type_validation)
            
            # Add spacing between sections
            current_row += 2
    
    # Now handle KAM pairs separately (exclude cases where salesperson = KAM)
    kam_pairs_query = text("""
        SELECT DISTINCT 
            d.SalespersonECode,
            sp.Name as SalespersonName,
            d.KAMECode,
            kam.Name as KAMName,
            a.Name as AccountName,
            a.Division,
            a.Location
        FROM Deals d
        JOIN Accounts a ON d.AccountID = a.ID
        JOIN Users sp ON d.SalespersonECode = sp.ECode
        JOIN Users kam ON d.KAMECode = kam.ECode
        WHERE d.Stage = 'DEAL_WON'
        AND d.KAMECode IS NOT NULL
        AND d.SalespersonECode != d.KAMECode
        ORDER BY sp.Name, kam.Name, a.Name
    """)
    
    kam_pairs_result = db.execute(kam_pairs_query, {"fy": fy, "month": month})
    kam_pairs = kam_pairs_result.fetchall()
    
    # Group KAM pairs by (SalespersonECode, KAMECode)
    kam_pair_groups = {}
    for row in kam_pairs:
        pair_key = (row.SalespersonECode, row.KAMECode)
        if pair_key not in kam_pair_groups:
            kam_pair_groups[pair_key] = {
                'salesperson_name': row.SalespersonName,
                'kam_name': row.KAMName,
                'accounts': []
            }
        kam_pair_groups[pair_key]['accounts'].append((row.AccountName, row.Division, row.Location))
    
    # Add KAM pair sections
    for (sp_ecode, kam_ecode), pair_info in kam_pair_groups.items():
        if pair_info['accounts']:  # Only show if there are accounts
            # KAM pair header
            pair_cell = ws.cell(row=current_row, column=1, 
                              value=f"{pair_info['salesperson_name']} & {pair_info['kam_name']} (ECode: {sp_ecode} & {kam_ecode})")
            pair_cell.font = kam_pair_font
            pair_cell.fill = kam_pair_fill
            ws.merge_cells(f"A{current_row}:D{current_row}")
            current_row += 1
            
            # Target info (combined target for the pair)
            target_cell = ws.cell(row=current_row, column=1, value=f"KAM Pair Business")
            target_cell.font = header_font
            ws.merge_cells(f"A{current_row}:D{current_row}")
            current_row += 1
            
            # Table headers
            table_headers = ["Account Name", "Customer Type", "Actuals", "Qualified"]
            for col_num, header in enumerate(table_headers, 1):
                cell = ws.cell(row=current_row, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            current_row += 1
            
            # Create dropdown validation for Customer Type column
            type_validation = DataValidation(
                type="list",
                formula1='"NEW,EXISTING"',
                allow_blank=True
            )
            
            # Add KAM pair account rows
            start_data_row = current_row
            for account_name, division, location in pair_info['accounts']:
                # Format account name
                formatted_name = account_name
                if division:
                    formatted_name += f" - {division}"
                if location:
                    formatted_name += f" - {location}"
                
                ws.cell(row=current_row, column=1, value=formatted_name).border = border
                type_cell = ws.cell(row=current_row, column=2, value="")
                type_cell.border = border
                ws.cell(row=current_row, column=3, value="").border = border
                ws.cell(row=current_row, column=4, value="").border = border
                current_row += 1
            
            # Apply dropdown validation to TYPE column for this pair's section
            end_data_row = current_row - 1
            if end_data_row >= start_data_row:
                type_validation.add(f"B{start_data_row}:B{end_data_row}")
                ws.add_data_validation(type_validation)
            
            # Add spacing between pairs
            current_row += 2
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
